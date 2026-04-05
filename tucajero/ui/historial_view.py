from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QDateEdit,
    QComboBox,
    QMessageBox,
    QInputDialog,
    QDialog,
)
from PySide6.QtCore import Qt, QDate
from datetime import datetime
from tucajero.ui.design_tokens import Colors, Typography, Spacing, BorderRadius
from tucajero.ui.components_premium import ButtonPremium, TABLE_STYLE_PREMIUM
import os
from tucajero.utils.formato import fmt_moneda


class HistorialView(QWidget):
    """Vista de historial de cierres y ranking de productos"""

    def __init__(self, session, parent=None):
        super().__init__(parent)
        self.session = session
        self.init_ui()
        self.cargar_historial()

    def init_ui(self):
        """Inicializa la interfaz"""
        self.setStyleSheet(f"background-color: {Colors.BG_APP};")

        layout = QVBoxLayout()
        self.setLayout(layout)

        titulo = QLabel("Historial de Cierres")
        titulo.setStyleSheet(
            f"font-size: {Typography.H2}px; font-weight: {Typography.BOLD}; color: {Colors.TEXT_PRIMARY};"
        )
        layout.addWidget(titulo)

        filtros_layout = QHBoxLayout()

        filtros_layout.addWidget(QLabel("Desde:"))
        self.desde_edit = QDateEdit()
        self.desde_edit.setCalendarPopup(True)
        self.desde_edit.setDate(QDate.currentDate().addDays(-30))
        filtros_layout.addWidget(self.desde_edit)

        filtros_layout.addWidget(QLabel("Hasta:"))
        self.hasta_edit = QDateEdit()
        self.hasta_edit.setCalendarPopup(True)
        self.hasta_edit.setDate(QDate.currentDate())
        filtros_layout.addWidget(self.hasta_edit)

        self.filtro_rapido = QComboBox()
        self.filtro_rapido.addItems(
            [
                "Filtro rápido...",
                "Este mes",
                "Mes anterior",
                "Últimos 3 meses",
                "Este año",
            ]
        )
        self.filtro_rapido.currentTextChanged.connect(self.on_filtro_rapido_changed)
        filtros_layout.addWidget(self.filtro_rapido)

        btn_filtrar = ButtonPremium("Filtrar", style="primary")
        btn_filtrar.clicked.connect(self.cargar_historial)
        filtros_layout.addWidget(btn_filtrar)

        btn_exportar = ButtonPremium("Exportar Excel", style="primary")
        btn_exportar.clicked.connect(self.exportar_excel)
        filtros_layout.addWidget(btn_exportar)

        btn_reimprimir = ButtonPremium("🖨 Reimprimir Ticket", style="secondary")
        btn_reimprimir.clicked.connect(self.reimprimir_ticket)
        filtros_layout.addWidget(btn_reimprimir)

        filtros_layout.addStretch()
        layout.addLayout(filtros_layout)

        self.resumen_widget = QWidget()
        self.resumen_widget.setStyleSheet(f"""
            QWidget {{
                background-color: {Colors.BG_ELEVATED};
                border-radius: {BorderRadius.MD}px;
                padding: {Spacing.MD}px;
            }}
        """)
        resumen_layout = QHBoxLayout()
        self.resumen_widget.setLayout(resumen_layout)

        self.lbl_ventas = QLabel(f"Ventas brutas: {fmt_moneda(0)}")
        self.lbl_ventas.setStyleSheet(
            f"color: {Colors.TEXT_PRIMARY}; font-size: {Typography.H5}px; font-weight: {Typography.BOLD};"
        )
        resumen_layout.addWidget(self.lbl_ventas)

        self.lbl_gastos = QLabel(f"Gastos: {fmt_moneda(0)}")
        self.lbl_gastos.setStyleSheet(
            f"color: {Colors.DANGER}; font-size: {Typography.H5}px; font-weight: {Typography.BOLD};"
        )
        resumen_layout.addWidget(self.lbl_gastos)

        self.lbl_ganancia = QLabel(f"Ganancia neta: {fmt_moneda(0)}")
        self.lbl_ganancia.setStyleSheet(
            f"color: {Colors.SUCCESS}; font-size: {Typography.H5}px; font-weight: {Typography.BOLD};"
        )
        resumen_layout.addWidget(self.lbl_ganancia)

        resumen_layout.addStretch()

        self.lbl_resumen_datos = QLabel("Cierres: 0 | Ventas: 0")
        self.lbl_resumen_datos.setStyleSheet(f"color: {Colors.TEXT_MUTED};")
        resumen_layout.addWidget(self.lbl_resumen_datos)

        layout.addWidget(self.resumen_widget)

        cierres_label = QLabel("Cierres del período")
        cierres_label.setStyleSheet(
            f"font-size: {Typography.H4}px; font-weight: {Typography.BOLD}; margin-top: {Spacing.SM}px; color: {Colors.TEXT_PRIMARY};"
        )
        layout.addWidget(cierres_label)

        self.tabla_cierres = QTableWidget()
        self.tabla_cierres.setColumnCount(5)
        self.tabla_cierres.setHorizontalHeaderLabels(
            ["ID", "Apertura", "Cierre", "Ventas brutas", "Ganancia neta"]
        )
        self.tabla_cierres.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self.tabla_cierres.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla_cierres.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla_cierres.setStyleSheet(TABLE_STYLE_PREMIUM)
        self.tabla_cierres.setMinimumHeight(200)
        layout.addWidget(self.tabla_cierres)

        ranking_label = QLabel("Ranking de productos")
        ranking_label.setStyleSheet(
            f"font-size: {Typography.H4}px; font-weight: {Typography.BOLD}; margin-top: {Spacing.SM}px;"
        )
        layout.addWidget(ranking_label)

        self.tabla_ranking = QTableWidget()
        self.tabla_ranking.setColumnCount(4)
        self.tabla_ranking.setHorizontalHeaderLabels(
            ["#", "Producto", "Unidades vendidas", "Ingresos"]
        )
        self.tabla_ranking.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self.tabla_ranking.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla_ranking.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla_ranking.setStyleSheet(TABLE_STYLE_PREMIUM)
        self.tabla_ranking.setMaximumHeight(250)
        layout.addWidget(self.tabla_ranking)

    def on_filtro_rapido_changed(self, text):
        """Aplica filtro rápido"""
        today = QDate.currentDate()
        if text == "Este mes":
            self.desde_edit.setDate(QDate(today.year(), today.month(), 1))
            self.hasta_edit.setDate(today)
        elif text == "Mes anterior":
            first_this_month = QDate(today.year(), today.month(), 1)
            last_prev_month = first_this_month.addDays(-1)
            first_prev_month = QDate(last_prev_month.year(), last_prev_month.month(), 1)
            self.desde_edit.setDate(first_prev_month)
            self.hasta_edit.setDate(last_prev_month)
        elif text == "Últimos 3 meses":
            self.desde_edit.setDate(today.addDays(-90))
            self.hasta_edit.setDate(today)
        elif text == "Este año":
            self.desde_edit.setDate(QDate(today.year(), 1, 1))
            self.hasta_edit.setDate(today)

    def cargar_historial(self):
        """Carga el historial de cierres"""
        desde = self.desde_edit.date().toPython()
        hasta = self.hasta_edit.date().toPython()
        hasta = datetime.combine(hasta, datetime.max.time())

        from tucajero.models.producto import CorteCaja, VentaItem, Venta, Producto
        from sqlalchemy import and_

        cierres = (
            self.session.query(CorteCaja)
            .filter(
                CorteCaja.fecha_cierre.isnot(None),
                CorteCaja.fecha_apertura
                >= datetime.combine(desde, datetime.min.time()),
                CorteCaja.fecha_apertura <= hasta,
            )
            .order_by(CorteCaja.fecha_apertura.desc())
            .all()
        )

        self.tabla_cierres.setRowCount(len(cierres))

        total_ventas = 0
        total_gastos = 0
        total_cierres = len(cierres)
        total_num_ventas = 0

        for i, corte in enumerate(cierres):
            self.tabla_cierres.setItem(i, 0, QTableWidgetItem(str(corte.id)))
            self.tabla_cierres.setItem(
                i,
                1,
                QTableWidgetItem(corte.fecha_apertura.strftime("%d/%m/%Y %I:%M %p")),
            )
            self.tabla_cierres.setItem(
                i,
                2,
                QTableWidgetItem(corte.fecha_cierre.strftime("%d/%m/%Y %I:%M %p"))
                if corte.fecha_cierre
                else QTableWidgetItem("—"),
            )
            self.tabla_cierres.setItem(
                i, 3, QTableWidgetItem(fmt_moneda(corte.total_ventas))
            )
            ganancia = getattr(corte, "ganancia_neta", corte.total_ventas)
            self.tabla_cierres.setItem(i, 4, QTableWidgetItem(fmt_moneda(ganancia)))

            total_ventas += corte.total_ventas
            total_gastos += getattr(corte, "total_gastos", 0)
            total_num_ventas += corte.numero_ventas

        self.lbl_ventas.setText(f"Ventas brutas: {fmt_moneda(total_ventas)}")
        self.lbl_gastos.setText(f"Gastos: {fmt_moneda(total_gastos)}")
        self.lbl_ganancia.setText(
            f"Ganancia neta: {fmt_moneda(total_ventas - total_gastos)}"
        )
        self.lbl_resumen_datos.setText(
            f"Cierres: {total_cierres} | Ventas: {total_num_ventas}"
        )

        self.cargar_ranking(desde, hasta)

    def cargar_ranking(self, desde, hasta):
        """Carga el ranking de productos"""
        desde_dt = datetime.combine(desde, datetime.min.time())
        hasta_dt = datetime.combine(hasta, datetime.max.time())

        from tucajero.models.producto import VentaItem, Venta, Producto
        from sqlalchemy import func

        ranking = (
            self.session.query(
                Producto.nombre,
                func.sum(VentaItem.cantidad).label("unidades"),
                func.sum(VentaItem.cantidad * VentaItem.precio).label("ingresos"),
            )
            .join(VentaItem, VentaItem.producto_id == Producto.id)
            .join(Venta, Venta.id == VentaItem.venta_id)
            .filter(
                Venta.fecha >= desde_dt,
                Venta.fecha <= hasta_dt,
                Venta.anulada == False,
            )
            .group_by(Producto.id, Producto.nombre)
            .order_by(func.sum(VentaItem.cantidad * VentaItem.precio).desc())
            .limit(50)
            .all()
        )

        self.tabla_ranking.setRowCount(len(ranking))

        for i, item in enumerate(ranking):
            self.tabla_ranking.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.tabla_ranking.setItem(i, 1, QTableWidgetItem(item.nombre))
            self.tabla_ranking.setItem(i, 2, QTableWidgetItem(str(int(item.unidades))))
            self.tabla_ranking.setItem(
                i, 3, QTableWidgetItem(fmt_moneda(item.ingresos))
            )

    def exportar_excel(self):
        """Exporta el historial a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()

            desde = self.desde_edit.date().toPython()
            hasta = self.hasta_edit.date().toPython()
            fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")

            downloads = os.path.join(os.path.expanduser("~"), "Downloads")
            filename = os.path.join(downloads, f"TuCajero_Historial_{fecha_str}.xlsx")

            ws_cierres = wb.active
            ws_cierres.title = "Cierres"

            headers = [
                "ID",
                "Apertura",
                "Cierre",
                "Ventas brutas",
                "Gastos",
                "Ganancia neta",
                "N° Ventas",
            ]
            ws_cierres.append(headers)

            for cell in ws_cierres[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="3498db", end_color="3498db", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            for row in range(self.tabla_cierres.rowCount()):
                row_data = []
                for col in range(self.tabla_cierres.columnCount()):
                    item = self.tabla_cierres.item(row, col)
                    row_data.append(item.text() if item else "")
                total = (
                    self.tabla_cierres.item(row, 3).text()
                    if self.tabla_cierres.item(row, 3)
                    else "$0.00"
                )
                row_data.extend(["$0.00", "$0.00", "0"])
                ws_cierres.append(row_data)

            ws_ranking = wb.create_sheet("Ranking")
            ranking_headers = ["#", "Producto", "Unidades vendidas", "Ingresos"]
            ws_ranking.append(ranking_headers)

            for cell in ws_ranking[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="27ae60", end_color="27ae60", fill_type="solid"
                )

            for row in range(self.tabla_ranking.rowCount()):
                row_data = []
                for col in range(self.tabla_ranking.columnCount()):
                    item = self.tabla_ranking.item(row, col)
                    row_data.append(item.text() if item else "")
                ws_ranking.append(row_data)

            for ws in [ws_cierres, ws_ranking]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            wb.save(filename)
            QMessageBox.information(
                self,
                "Exportación exitosa",
                f"Archivo exportado a:\n{filename}",
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error al exportar",
                f"No se pudo exportar el archivo:\n{str(e)}",
            )

    def reimprimir_ticket(self):
        """Reimprime el ticket de una venta por su ID"""
        venta_id, ok = QInputDialog.getInt(
            self,
            "Reimprimir Ticket",
            "Ingrese el número de ticket a reimprimir:",
            1,  # default
            1,  # min
            999999,  # max
            1,  # step
        )
        if not ok:
            return

        try:
            from tucajero.models.producto import Venta, VentaItem
            from tucajero.utils.ticket import GeneradorTicket

            venta = (
                self.session.query(Venta)
                .filter(Venta.id == venta_id)
                .first()
            )
            if not venta:
                QMessageBox.warning(
                    self,
                    "Ticket no encontrado",
                    f"No se encontró la venta #{venta_id}.",
                )
                return

            if venta.anulada:
                QMessageBox.warning(
                    self,
                    "Venta anulada",
                    f"La venta #{venta_id} está anulada.\n"
                    f"Motivo: {getattr(venta, 'motivo_anulacion', 'No especificado')}",
                )
                return

            items = (
                self.session.query(VentaItem)
                .filter(VentaItem.venta_id == venta.id)
                .all()
            )
            if not items:
                QMessageBox.warning(
                    self,
                    "Sin detalle",
                    f"La venta #{venta_id} no tiene productos registrados.",
                )
                return

            generador = GeneradorTicket()
            ticket_text = generador.generar(venta, items)

            # Mostrar preview en dialog
            from PySide6.QtWidgets import QTextEdit
            preview = QDialog(self)
            preview.setWindowTitle("Vista previa del ticket")
            preview.setMinimumSize(400, 600)
            preview_layout = QVBoxLayout(preview)

            text_edit = QTextEdit()
            text_edit.setPlainText(ticket_text)
            text_edit.setReadOnly(True)
            text_edit.setStyleSheet(f"""
                QTextEdit {{
                    background-color: {Colors.BG_INPUT};
                    color: {Colors.TEXT_PRIMARY};
                    border: 1px solid {Colors.BORDER_DEFAULT};
                    border-radius: {BorderRadius.MD}px;
                    padding: {Spacing.MD}px;
                    font-family: 'Consolas', 'Courier New', monospace;
                    font-size: 12px;
                }}
            """)
            preview_layout.addWidget(text_edit)

            btn_layout = QHBoxLayout()
            btn_print = ButtonPremium("🖨 Imprimir", style="primary")
            btn_print.clicked.connect(lambda: self._imprimir_venta(venta, items, preview))
            btn_layout.addWidget(btn_print)

            btn_email = ButtonPremium("📧 Enviar por Email", style="secondary")
            btn_email.clicked.connect(lambda: self._enviar_email_venta(venta, items, preview))
            btn_layout.addWidget(btn_email)

            btn_close = ButtonPremium("Cerrar", style="secondary")
            btn_close.clicked.connect(preview.reject)
            btn_layout.addWidget(btn_close)

            preview_layout.addLayout(btn_layout)
            preview.exec()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error al reimprimir",
                f"No se pudo reimprimir el ticket:\n{str(e)}",
            )

    def _imprimir_venta(self, venta, items, parent_dialog=None):
        """Imprime el ticket usando la impresora térmica"""
        try:
            from tucajero.utils.ticket import GeneradorTicket
            from tucajero.utils.store_config import get_printer_enabled
            from tucajero.utils.impresora import get_impresora

            generador = GeneradorTicket()
            generador.imprimir(venta, items)

            if get_printer_enabled():
                imp = get_impresora()
                imp.imprimir_ticket(venta, items)
                imp.desconectar()

            # Registrar en auditoría
            try:
                from tucajero.services.audit_service import AuditService
                audit = AuditService(self.session)
                audit.registrar(
                    AuditService.REIMPRESION,
                    f"Reimpresión de ticket #{venta.id}",
                    entidad_tipo="Venta",
                    entidad_id=venta.id,
                )
            except Exception:
                pass

            QMessageBox.information(
                self,
                "Impresión exitosa",
                f"Ticket #{venta.id} enviado a imprimir.",
            )
            if parent_dialog:
                parent_dialog.accept()
        except Exception as e:
            QMessageBox.warning(
                self,
                "Error de impresión",
                f"No se pudo imprimir en la impresora térmica:\n{str(e)}\n\n"
                f"El ticket se guardó en el PDF diario.",
            )

    def _enviar_email_venta(self, venta, items, parent_dialog=None):
        """Envía el ticket por email al cliente"""
        # Intentar obtener email del cliente asociado
        email_cliente = None
        try:
            if hasattr(venta, "cliente_id") and venta.cliente_id:
                from tucajero.models.cliente import Cliente
                cliente = (
                    self.session.query(Cliente)
                    .filter(Cliente.id == venta.cliente_id)
                    .first()
                )
                if cliente and cliente.email:
                    email_cliente = cliente.email
        except Exception:
            pass

        # Pedir email al usuario
        email, ok = QInputDialog.getText(
            self,
            "Enviar por Email",
            "Email del destinatario:",
            text=email_cliente or "",
        )
        if not ok or not email.strip():
            return

        email = email.strip()

        # Confirmar envío
        respuesta = QMessageBox.question(
            self,
            "Confirmar envío",
            f"¿Enviar ticket #{venta.id} a {email}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if respuesta != QMessageBox.StandardButton.Yes:
            return

        try:
            from tucajero.utils.email_envio import enviar_ticket_email

            success, message = enviar_ticket_email(email, venta, items)
            if success:
                # Registrar en auditoría
                try:
                    from tucajero.services.audit_service import AuditService
                    audit = AuditService(self.session)
                    audit.registrar(
                        AuditService.EMAIL_TICKET,
                        f"Ticket #{venta.id} enviado a {email}",
                        entidad_tipo="Venta",
                        entidad_id=venta.id,
                    )
                except Exception:
                    pass
                QMessageBox.information(self, "Email enviado", message)
                if parent_dialog:
                    parent_dialog.accept()
            else:
                QMessageBox.warning(self, "Error al enviar", message)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error al enviar email",
                f"No se pudo enviar el ticket:\n{str(e)}",
            )
