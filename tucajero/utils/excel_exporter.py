import os
from datetime import datetime


def exportar_historial_excel(
    cierres, ventas_por_cierre, ranking, resumen, ruta_destino=None
):
    """Exporta historial a Excel con 3 hojas"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Cierres"
    headers1 = [
        "ID",
        "Fecha apertura",
        "Fecha cierre",
        "Ventas brutas",
        "IVA recaudado",
        "Efectivo",
        "Transferencias",
        "Gastos",
        "Ganancia neta",
        "Núm. ventas",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2C3E50")
        cell.font = Font(bold=True, color="FFFFFF")
    for c in cierres:
        ws1.append(
            [
                c.id,
                c.fecha_apertura.strftime("%d/%m/%Y %H:%M") if c.fecha_apertura else "",
                c.fecha_cierre.strftime("%d/%m/%Y %H:%M") if c.fecha_cierre else "",
                round(c.total_ventas or 0, 2),
                round(c.total_iva or 0, 2),
                round(c.total_efectivo or 0, 2),
                round(c.total_transferencias or 0, 2),
                round(c.total_gastos or 0, 2),
                round(c.ganancia_neta or 0, 2),
                c.numero_ventas or 0,
            ]
        )
    ws1.append(
        [
            "TOTALES",
            "",
            "",
            round(resumen["total_ventas"], 2),
            round(sum(c.total_iva or 0 for c in cierres), 2),
            round(sum(c.total_efectivo or 0 for c in cierres), 2),
            round(sum(c.total_transferencias or 0 for c in cierres), 2),
            round(resumen["total_gastos"], 2),
            round(resumen["ganancia_neta"], 2),
            resumen["num_ventas"],
        ]
    )
    for cell in ws1[ws1.max_row]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("Detalle ventas")
    headers2 = [
        "Cierre ID",
        "Venta ID",
        "Fecha",
        "Subtotal base",
        "IVA 19%",
        "Total",
        "Método de pago",
        "Estado",
        "Motivo anulación",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="27AE60")
        cell.font = Font(bold=True, color="FFFFFF")
    IVA_RATE = 0.19
    for corte_id, ventas in ventas_por_cierre.items():
        for v in ventas:
            subtotal = round(v.total / 1.19, 2)
            iva = round(v.total - subtotal, 2)
            ws2.append(
                [
                    corte_id,
                    v.id,
                    v.fecha.strftime("%d/%m/%Y %H:%M:%S"),
                    subtotal,
                    iva,
                    round(v.total, 2),
                    v.metodo_pago or "efectivo",
                    "Anulada" if v.anulada else "Válida",
                    v.motivo_anulacion or "",
                ]
            )

    ws3 = wb.create_sheet("Ranking productos")
    headers3 = [
        "Posición",
        "Código",
        "Producto",
        "Unidades vendidas",
        "Ingresos totales",
    ]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="8E44AD")
        cell.font = Font(bold=True, color="FFFFFF")
    for i, r in enumerate(ranking, 1):
        ws3.append(
            [
                i,
                r.codigo,
                r.nombre,
                int(r.total_vendido or 0),
                round(float(r.total_ingresos or 0), 2),
            ]
        )

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    if not ruta_destino:
        from config.database import get_data_dir

        exports_dir = os.path.join(get_data_dir(), "exports")
        os.makedirs(exports_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_destino = os.path.join(exports_dir, f"historial_{timestamp}.xlsx")

    wb.save(ruta_destino)
    return ruta_destino
