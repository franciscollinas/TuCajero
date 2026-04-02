import openpyxl
import csv
import os
import re
import unicodedata

HEADER_MAP = {
    "codigo": "codigo",
    "code": "codigo",
    "cod": "codigo",
    "codigo_producto": "codigo",
    "codigo_de_barras": "codigo",
    "barcode": "codigo",
    "sku": "codigo",
    "ref": "codigo",
    "nombre": "nombre",
    "name": "nombre",
    "producto": "nombre",
    "descripcion": "nombre",
    "description": "nombre",
    "nombre_producto": "nombre",
    "nombre_del_producto": "nombre",
    "articulo": "nombre",
    "articulo_nombre": "nombre",
    "precio": "precio",
    "price": "precio",
    "precio_venta": "precio",
    "precio_de_venta": "precio",
    "precio_ventas": "precio",
    "pvp": "precio",
    "valor": "precio",
    "venta": "precio",
    "valor_unitario": "precio",
    "precio_unitario": "precio",
    "costo": "costo",
    "cost": "costo",
    "precio_costo": "costo",
    "costo_compra": "costo",
    "precio_de_compra": "costo",
    "compra": "costo",
    "costo_unitario": "costo",
    "stock": "stock",
    "cantidad": "stock",
    "inventory": "stock",
    "existencia": "stock",
    "existencias": "stock",
    "qty": "stock",
    "unidades": "stock",
    "stock_disponible": "stock",
    "cantidad_disponible": "stock",
    "categoria": "categoria",
    "category": "categoria",
    "tipo": "categoria",
    "grupo": "categoria",
    "linea": "categoria",
    "departamento": "categoria",
    "familia": "categoria",
    "aplica_iva": "aplica_iva",
    "iva": "aplica_iva",
    "aplica_iva_19": "aplica_iva",
    "con_iva": "aplica_iva",
    "gravado": "aplica_iva",
    "taxable": "aplica_iva",
    "aplica_i_v_a": "aplica_iva",
}


def normalizar_header(h):
    if not h:
        return ""
    h = str(h).strip().lower()
    h = unicodedata.normalize("NFD", h)
    h = "".join(c for c in h if unicodedata.category(c) != "Mn")
    h = re.sub(r"[^a-z0-9]+", "_", h)
    h = h.strip("_")
    return h


def mapear_headers(headers_raw):
    resultado = {}
    col_map = {normalizar_header(c): c for c in headers_raw}

    for i, h in enumerate(headers_raw):
        norm = normalizar_header(h)
        estandar = HEADER_MAP.get(norm)

        if not estandar:
            for variante, std in HEADER_MAP.items():
                if variante in norm or norm in variante:
                    estandar = std
                    break
                if std not in resultado.values():
                    words = norm.split("_")
                    if all(w in variante for w in words if len(w) > 2):
                        estandar = std
                        break

        if estandar:
            resultado[i] = estandar
        else:
            resultado[i] = norm

    return resultado


def validar_columnas_requeridas(filas):
    required = ["codigo", "nombre", "precio"]
    if not filas:
        return []

    columnas_disponibles = set()
    for fila in filas[:5]:
        columnas_disponibles.update(fila.keys())

    faltantes = []
    for col in required:
        if col not in columnas_disponibles:
            faltantes.append(col)

    return faltantes


def leer_archivo(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    filas = []

    if ext in [".xlsx", ".xls"]:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers_raw = [c.value for c in ws[1]]
        mapa = mapear_headers(headers_raw)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                fila = {}
                for i, val in enumerate(row):
                    if i in mapa:
                        fila[mapa[i]] = val
                filas.append(fila)

    elif ext == ".csv":
        with open(filepath, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            headers_raw = next(reader, [])
            mapa = mapear_headers(headers_raw)
            for row in reader:
                if any(v.strip() for v in row):
                    fila = {}
                    for i, val in enumerate(row):
                        if i < len(mapa) and i in mapa:
                            fila[mapa[i]] = val
                    filas.append(fila)

    return filas


def importar_productos(filepath, session):
    from tucajero.models.producto import Producto, Categoria

    filas = leer_archivo(filepath)

    faltantes = validar_columnas_requeridas(filas)
    if faltantes:
        return {
            "importados": 0,
            "actualizados": 0,
            "errores": [
                {
                    "fila": 1,
                    "msg": f"Columnas requeridas faltantes: {', '.join(faltantes)}",
                }
            ],
        }

    importados = actualizados = 0
    errores = []
    categorias_cache = {}

    for i, fila in enumerate(filas, start=2):
        try:
            codigo = str(fila.get("codigo", "") or "").strip()
            nombre = str(fila.get("nombre", "") or "").strip()
            if not codigo or not nombre:
                errores.append({"fila": i, "msg": "Código y nombre requeridos"})
                continue

            precio = float(str(fila.get("precio", 0) or 0).replace(",", "."))
            costo = float(str(fila.get("costo", 0) or 0).replace(",", "."))
            stock = int(float(str(fila.get("stock", 0) or 0)))
            iva_raw = str(fila.get("aplica_iva", "SI") or "SI").upper().strip()
            aplica_iva = iva_raw not in ["NO", "FALSE", "0", "N"]

            cat_nombre = str(fila.get("categoria", "") or "").strip()
            categoria_id = None
            if cat_nombre:
                if cat_nombre not in categorias_cache:
                    cat = session.query(Categoria).filter_by(nombre=cat_nombre).first()
                    if not cat:
                        cat = Categoria(nombre=cat_nombre)
                        session.add(cat)
                        session.flush()
                    categorias_cache[cat_nombre] = cat.id
                categoria_id = categorias_cache[cat_nombre]

            existente = session.query(Producto).filter_by(codigo=codigo).first()
            if existente:
                existente.nombre = nombre
                existente.precio = precio
                existente.costo = costo
                existente.stock = stock
                existente.aplica_iva = aplica_iva
                existente.categoria_id = categoria_id
                existente.activo = True
                actualizados += 1
            else:
                session.add(
                    Producto(
                        codigo=codigo,
                        nombre=nombre,
                        precio=precio,
                        costo=costo,
                        stock=stock,
                        aplica_iva=aplica_iva,
                        categoria_id=categoria_id,
                    )
                )
                importados += 1
        except Exception as e:
            errores.append({"fila": i, "msg": str(e)})

    session.commit()
    return {
        "importados": importados,
        "actualizados": actualizados,
        "errores": errores,
    }
